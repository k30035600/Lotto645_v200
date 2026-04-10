# -*- coding: utf-8 -*-
"""
Lotto645.xlsx 검증 및 보완 스크립트.
1회부터 최신 추첨 회차까지 검증하고, 빠진 회차는 get_lotto_round 모듈로 API 조회 후 Excel에 추가합니다.
사용: python sync_lotto645_xlsx.py
필요: openpyxl (pip install openpyxl)
"""
import sys
import os
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except (AttributeError, OSError):
    pass

try:
    import openpyxl
except ImportError:
    print('openpyxl이 필요합니다. 실행: pip install openpyxl')
    sys.exit(1)

# 프로젝트 루트 기준
BASE_DIR = Path(__file__).resolve().parent.parent
XLSX_PATH = BASE_DIR / 'source' / 'Lotto645.xlsx'

# app.js와 동일한 컬럼명 (Lotto645.xlsx)
HEADERS = ['회차', '날짜', '번호1', '번호2', '번호3', '번호4', '번호5', '번호6', '보너스번호']


def main():
    xlsx_abs = XLSX_PATH.resolve()
    print('[대상] Lotto645.xlsx 경로: %s' % xlsx_abs)
    if not XLSX_PATH.is_file():
        print('[오류] 파일이 없습니다: %s' % xlsx_abs)
        sys.exit(1)

    # 상위 디렉터리를 sys.path에 추가하여 utils 패키지를 찾을 수 있게 함
    sys.path.append(str(Path(__file__).resolve().parent.parent))
    from utils.get_lotto_round import get_latest_round_no, get_lotto_number

    latest = get_latest_round_no()
    if latest is None:
        print('[오류] 최신 회차를 API에서 가져오지 못했습니다.')
        sys.exit(1)
    print('[동행복권 최신 추첨정보] 시작회차: 1회, 마지막회차: %s회 (당첨회차)' % latest)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    if ws.max_row < 2:
        print('[오류] Lotto645.xlsx에 헤더만 있거나 비어 있습니다.')
        sys.exit(1)

    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_row = [str(h).strip() if h is not None else '' for h in header_row]
    idx_round = idx_date = idx_bonus = None
    idx_nums = []
    for i, h in enumerate(header_row):
        c = i + 1
        if h == '회차':
            idx_round = c
        elif h == '날짜':
            idx_date = c
        elif h == '보너스번호':
            idx_bonus = c
        elif h and len(h) == 2 and h.startswith('번호') and h[1].isdigit():
            idx_nums.append((int(h[1]), c))
    idx_nums.sort(key=lambda x: x[0])
    idx_nums = [col for _, col in idx_nums]
    if idx_round is None or idx_date is None or len(idx_nums) != 6 or idx_bonus is None:
        if header_row[:9] == HEADERS:
            idx_round, idx_date = 1, 2
            idx_nums = list(range(3, 9))
            idx_bonus = 9
        else:
            print('[오류] 필요한 헤더(회차, 날짜, 번호1~6, 보너스번호)를 찾을 수 없습니다.')
            sys.exit(1)

    existing_rounds = set()
    rows_by_round = {}
    for r in range(2, ws.max_row + 1):
        rno_val = ws.cell(r, idx_round).value
        try:
            rno = int(rno_val)
        except (TypeError, ValueError):
            continue
        existing_rounds.add(rno)
        row_vals = [ws.cell(r, col).value for col in range(1, ws.max_column + 1)]
        rows_by_round[rno] = row_vals

    need_rounds = sorted(set(range(1, latest + 1)) - existing_rounds)
    if not need_rounds:
        print('[완료] 1회 ~ %s회까지 누락된 회차가 없습니다.' % latest)
        return

    # 빠진 회차정보 콘솔 출력
    print('')
    print('[빠진 회차정보] %s건' % len(need_rounds))
    print('  회차 목록: %s' % (', '.join('%s회' % r for r in need_rounds)))
    for rno in need_rounds:
        print('    - %s회' % rno)
    print('')

    added = 0
    for rno in need_rounds:
        result = get_lotto_number(rno)
        if result is None:
            print('[경고] %s회 조회 실패, 건너뜀.' % rno)
            continue
        # 새 행: 회차, 날짜, 번호1~6, 보너스번호 순 (기존 시트 컬럼 수에 맞춤)
        row_vals = [None] * ws.max_column
        row_vals[idx_round - 1] = result.get('drwNo')
        row_vals[idx_date - 1] = result.get('date') or ''
        nums = result.get('numbers', [])
        for i, col in enumerate(idx_nums):
            row_vals[col - 1] = nums[i] if i < len(nums) else None
        row_vals[idx_bonus - 1] = result.get('bonus')
        rows_by_round[rno] = row_vals
        added += 1
        print('  추가: %s회 (%s)' % (rno, result.get('date')))

    # 회차별 당첨번호 다시 생성 (최신회차 ~ 1회 내림차순)
    sorted_rounds = sorted(rows_by_round.keys(), reverse=True)
    print('')
    print('[회차별 당첨번호] 최신회차 ~ 1회 내림차순으로 Lotto645.xlsx에 다시 생성합니다.')
    # 기존 데이터 행 제거 (2행부터)
    while ws.max_row >= 2:
        ws.delete_rows(2, 1)
    for row_idx, rno in enumerate(sorted_rounds, start=2):
        row_vals = rows_by_round[rno]
        for col_idx, val in enumerate(row_vals, start=1):
            ws.cell(row_idx, col_idx, value=val)

    # Lotto645.xlsx가 수정되지 않는 경우 가능한 원인 (저장 전 안내)
    print('')
    print('[Lotto645.xlsx가 수정되지 않는 경우 가능한 원인]')
    print('  1. Excel에서 Lotto645.xlsx를 열어 둔 상태 (파일 잠김)')
    print('  2. OneDrive 동기화 중 파일 잠금 또는 충돌')
    print('  3. 다른 프로그램이 해당 파일 사용 중')
    print('  4. 파일이 읽기 전용이거나 쓰기 권한 없음')
    print('  5. 스크립트 실행 경로가 프로젝트 루트가 아님 (다른 파일을 수정함)')
    print('  → 해결: Excel/프로그램에서 파일을 닫고, 프로젝트 폴더에서 다시 실행')
    print('')

    try:
        wb.save(XLSX_PATH)
        print('[저장] %s — %s회 추가됨 (총 %s회 반영).' % (xlsx_abs, added, len(sorted_rounds)))
        print('[회차별 당첨번호] Lotto645.xlsx에 1회 ~ %s회 반영 완료. 브라우저에서 새로고침하면 화면에 반영됩니다.' % latest)
    except PermissionError as e:
        print('[오류] 저장 실패(파일 잠김): %s' % e)
        print('[Lotto645.xlsx가 수정되지 않는 이유] Excel에서 파일이 열려 있거나 OneDrive가 파일을 잠갔을 수 있습니다.')
        print('  → Lotto645.xlsx를 Excel에서 닫고, OneDrive 동기화가 끝난 뒤 다시 실행하세요.')
        sys.exit(1)
    except Exception as e:
        print('[오류] 저장 실패: %s' % e)
        print('[Lotto645.xlsx가 수정되지 않는 이유] 위 오류를 확인한 뒤 파일을 닫고 다시 실행하거나, 관리자 권한으로 실행해 보세요.')
        sys.exit(1)


if __name__ == '__main__':
    main()
